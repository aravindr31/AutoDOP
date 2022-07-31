import re
import shutil
import sys

import pandas as pd
import openpyxl

from openpyxl.utils import get_column_interval
from copy import copy

filename = sys.argv[1]
addData = sys.argv[2]
basePath = 'C:\\Users\\Aravind\\Documents\\RD'
rFile = basePath + '\\xlsxFile'
fFile = basePath + '\\formatted'
parent = basePath + '\\Parent\\Custom_RD_Model.xlsx'


# dataframe function
def load_workbook_range(range_string, ws):
    col_start, col_end = re.findall("[A-Z]+", range_string)

    data_rows = []
    for row in ws[range_string]:
        data_rows.append([cell.value for cell in row])

    return pd.DataFrame(data_rows, columns=get_column_interval(col_start, col_end))


# insert Variable data function
def varDataIn(data, arrIndex):
    for row in range(14, 14 + len(dframe)):
        # for col in range(1, 12):
        for col, text in enumerate(data, start=2):
            ws.cell(column=col, row=14 + arrIndex, value=text)


# Formulate Dataframe
# wb = openpyxl.load_workbook(rFile + "\\RDInstallmentReport29-07-2022.xlsx")
wb = openpyxl.load_workbook(rFile + "\\" + filename)

wst = wb['RDInstallmentReport']

dframe = load_workbook_range('F14:U17', wst)
# constants

listCID = wst["I6"].value
createdDate = wst["W14"].value
fromDate = wst["I5"].value
totalAmt = wst["J" + str(14 + len(dframe) + 1)].value

wb.close()

# copy to formatted folder
copyChild = fFile + '\\' + listCID + '.xlsx'
shutil.copyfile(parent, fFile + '/' + listCID + '.xlsx')

formatedDframe = dframe.to_numpy().tolist()

for i in reversed(range(len(dframe))):
    for j in formatedDframe[len(dframe) - 1 - i][:]:
        if j is None:
            formatedDframe[len(dframe) - 1 - i].remove(j)

# Open Copied Child File
wb = openpyxl.load_workbook(copyChild)
ws = wb['Sheet1']
ws.insert_rows(14, len(dframe) - 2)

# insert Constant datas function

for x in range(len(dframe)):
    varDataIn(formatedDframe[x], x)


# insertConstant
def insertConst(data, col):
    for row in range(14, 14 + len(dframe)):
        ws.cell(column=col, row=row, value=data)


ws.cell(column=6, row=6, value=listCID)
ws.cell(column=6, row=5, value=fromDate)
ws.cell(column=6, row=11, value=totalAmt)
ws.cell(column=6, row=12, value=len(dframe))
ws.cell(column=6, row=14 + len(dframe) + 1, value=totalAmt)
ws.cell(column=1, row=14 + len(dframe) + 1, value=listCID)
ws.cell(column=8, row=1, value=addData)

insertConst(listCID, 1)
insertConst(createdDate, 11)
insertConst("Success", 10)
# merge
ws.merge_cells("A" + str(14 + len(dframe)) + ":E" + str(14 + len(dframe)))
ws.merge_cells("F" + str(14 + len(dframe)) + ":K" + str(14 + len(dframe)))
ws.merge_cells("A" + str(14 + len(dframe) + 1) + ":E" + str(14 + len(dframe) + 1))
ws.merge_cells("F" + str(14 + len(dframe) + 1) + ":K" + str(14 + len(dframe) + 1))

for row in ws.iter_rows(min_row=14, max_row=14 + len(dframe)):
    for cell in row:
        alignment_obj = copy(cell.alignment)
        alignment_obj.horizontal = 'left'
        alignment_obj.vertical = 'center'
        alignment_obj.wrapText = True
        cell.alignment = alignment_obj
#
# img = openpyxl.drawing.image.Image('postal.jpg')
# img.anchor = 'A1'
# ws.add_image(img)

for row in range(14, 14 + len(dframe)):
    ws.row_dimensions[row].height = 29

print("Saved in Downloads Folder " + listCID + ".xlsx")
wb.save(copyChild)
